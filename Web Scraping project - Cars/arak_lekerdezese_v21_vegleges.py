from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
from time import sleep
from tabula.io import read_pdf
import requests
import re
import pdfplumber
import io
from tqdm import tqdm


####################################################################################


#Skoda
def getSkodaPrices(SkodaModel, SkodaLink):
    
    driver.get(SkodaLink)

    # Sütik elfogadása
    try:
        accept_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@id="onetrust-accept-btn-handler"]'))
        )
        accept_button.click()
    except:
        None

    KonkretKonfigurator_element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.XPATH, f"//span[translate(normalize-space(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')=translate('{SkodaModel}', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')]")))
    KonkretKonfigurator_element.click()

    sleep(3)
    try:
        price_element = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, "//span[contains(@class, 'secondary ng-star-inserted')]"))
        )
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{clean_price}"
        print("Feldolgozás sikeres")
        return formatted_price
    
    except:
        print(f"Hiba történt ezzel a Skoda modellel: {SkodaModel}")



# Hyundai
def getHyundaiprices(HyundaiModel, HyundaiLink):
    Ár = 0

    df = pd.read_excel(ExcelFileBeolvasásra, header=0)
    Meghajtás = df["Meghajtás (Hyundai, Volvo)"].iloc[index,]

    driver.get(HyundaiLink)
    sleep(3)
    KeresettModelElement = driver.find_element(By.XPATH, f"//div[@class='name'][h3[translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')=translate('{HyundaiModel}', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')]]")
    Modell_div = KeresettModelElement.find_elements(By.XPATH, './../..')
    href_element = Modell_div[0].find_element(By.XPATH, ".//a[@href]")
    href_value = href_element.get_attribute('href')
    driver.get(href_value)

    ÁrlistaGomb = driver.find_element(By.XPATH, "//a[contains(text(), 'Árlista')]")
    ÁrlistaWeboldal = ÁrlistaGomb.get_attribute('href')
    driver.get(ÁrlistaWeboldal)

    ÁrlistaPdfGomb = driver.find_element(By.XPATH, "//div[contains(text(), 'Árlista')]")
    parent_link = ÁrlistaPdfGomb.find_element(By.XPATH, "..")
    ÁrlistaPdflink = parent_link.get_attribute('href')

    df = read_pdf(ÁrlistaPdflink, pages='all', multiple_tables=True)[0]

    try:
            # Azért dataframe 0. oszlopára szűrünk hogy megtaláljuk a keresett modellt, mert bármelyik másik Hyundai modell Árlista pdf-jében is az első oszlopban van a modell teljes neve
            # a copy nélkül a "dfSzűrt" nem egy önálló dataframe lenne, hanem csak egy referencia a "df-re", amivel így nem lehetne tovább dolgozni
            dfSzűrt = df[df.iloc[:, 0] == Meghajtás].copy()

            # A listaárakat átalakítjuk számmá, hogy utána ki tudjuk venni belőlük a legkisebbet
            dfSzűrt["Listaár"] = dfSzűrt["Listaár"].str.replace(" Ft", "").str.replace(" ", "").astype(int)

            # Legalacsonyabb listaár
            Ár = dfSzűrt["Listaár"].min()
            print("Feldolgozás sikeres")
    except:
            print(f"Hiba ennél a Hyundai modellnél | {{HyundaiModel}}:{HyundaiModel}, {{Ár}}: {Ár}")
    
    return Ár
    


# Kia
def getKiaPrices(KiaModel, KiaLink):
        KiaLink = KiaLink
        try:
            driver = webdriver.Chrome()
            driver.maximize_window()
            driver.get(KiaLink)

            formatted_price = None

            # Sütik elfogadása
            button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler")))
            button.click()

            # Megtaláljuk az adott modellt
            KonkretKonfigurator_element = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, f"//h3[normalize-space()='{KiaModel}']")))

            # Megkeressük a modellhez tartozó árat
            price_element = KonkretKonfigurator_element.find_element(By.XPATH, "following-sibling::span[contains(@class, 'price')]")
            price = price_element.text

            clean_price = ''.join(filter(str.isdigit, price))
            formatted_price = f"{int(clean_price)}"
            print("Feldolgozás sikeres")

        except:
            try:
                print(f"A következő Kia modellt újra be kellett tölteni {KiaModel}")
                driver = webdriver.Chrome()
                driver.maximize_window()
                driver.get(KiaLink)

                formatted_price = None

                # Sütik elfogadása
                button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler")))
                button.click()

                # Megtaláljuk az adott modellt
                KonkretKonfigurator_element = WebDriverWait(driver, 300).until(
                EC.visibility_of_element_located((By.XPATH, f"//h3[normalize-space()='{KiaModel}']")))

                # Megkeressük a modellhez tartozó árat
                price_element = KonkretKonfigurator_element.find_element(By.XPATH, "following-sibling::span[contains(@class, 'price')]")
                price = price_element.text

                clean_price = ''.join(filter(str.isdigit, price))
                formatted_price = f"{int(clean_price)}"
                print("Feldolgozás sikeres")

            except:
                 print(f"Hiba történt ezzel a Kia modellel: {KiaModel}")
                
        return formatted_price

#Audi
def getAudiPrices(AudiModel, AudiLink):
    driver.get(AudiLink)
    sleep(5)

    try:
        try:
            # Sütik elfogadása
            accept_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '//button[@id="onetrust-accept-btn-handler"]')))
            accept_button.click()
            sleep(3)
        
        except:
            pass
        
        first_space_index = AudiModel.find(' ')
        AudiModelMárkaNélkül = AudiModel[first_space_index+1:].lower()

        try:
            AudiModelFelbontva = AudiModelMárkaNélkül.split(" ") 
            FőModellNév = AudiModelFelbontva[0] # pl: AudiModelMárkaNélkül: "Q4 Sportback e-tron", FőModellNév: "Q4",
                                                # vagy pl: AudiModelMárkaNélkül: "Q3 Sportback/RSQ3 SB ", FőModellNév: "Q3"
            print(FőModellNév)

            KonkretKonfigurator_element = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, f'//div[@data-cy-modelclass="{FőModellNév}"]'))
            )
            KonkretKonfigurator_element.click()

        except:
                AudiModelFelbontva = AudiModelMárkaNélkül.split("/")
                FőModellNév = AudiModelFelbontva[0] # pl: AudiModelMárkaNélkül: "Q3/RSQ3", FőModellNév: "Q3"
                                                    # vagy pl. AudiModelMárkaNélkül: "Q5/Q5e Sportback/SQ5 Sportback", FőModellNév: "Q5"
                print(FőModellNév)

                KonkretKonfigurator_element = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, f'//div[@data-cy-modelclass="{FőModellNév}"]'))
                )
                KonkretKonfigurator_element.click()

        
        sleep(6)  # Idő adása az oldalbetöltésre
        konkret = driver.find_element(By.XPATH, f"//span[contains(text(), '{AudiModel}')][@class='group-name-inner']")
        konkret.click()
        sleep(10)

        # Az ár kinyerése
        price_element = driver.find_element(By.XPATH, '//span[@class="from"]')
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{int(clean_price)}"
        print("Feldolgozás sikeres")

    except:
        print(f"Hiba történt ezzel az Audi modellel: {AudiModel}")
        formatted_price = 0

    return formatted_price

#Peugeout
def getPeugeotPrices(PeugeotLink):
    formatted_price = 0

    driver.get(PeugeotLink)

    # Sütik elfogadása
    accept_button = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.ID, "_psaihm_id_accept_all_btn"))
        )
    # Elfogadás gombra kattintás
    accept_button.click()

    sleep(3)
    KonkretKonfigurator_element = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, f"(//span[@class='q-label'][normalize-space()='KONFIGURÁTOR'])[6]"))
            )
    KonkretKonfigurator_element.click()
    NavigaloGomb_element = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, f"//button[@id='cta-slider-next-selection-bar']//i[contains(@role,'presentation')]"))
            )
    NavigaloGomb_element.click()
    sleep(3)
    KonkretModell_element = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, f"/html/body/div[1]/div/div[2]/section/div/div[4]/div[1]/div/div/div/div/div/div/div/div/div/div[1]/div[1]/main/section/div[1]/div[1]/div[1]/div/div/div[5]/div/div/div"))
            )
    KonkretModell_element.click()
    sleep(3)
    price_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "(//span[@class='cash-price brand-p'])[1]"))
            )
    price_element = price_element.text
    clean_price = ''.join(filter(str.isdigit, price_element))
    formatted_price = f"{clean_price}"
    print("Feldolgozás sikeres")
    return formatted_price
    

#Suzuki
def getSuzukiPrices(SuzukiModel, SuzukiLink):
    driver.get(SuzukiLink)
    sleep(3)

    try:
        SuzukiModel = SuzukiModel.lower()
        models = driver.find_elements(By.XPATH, f"//div[h2[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{SuzukiModel}')]]")

        price_element = models[1].find_element(By.XPATH, ".//following::strong[1]")
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{int(clean_price)}"
        print("Feldolgozás sikeres")
    except:
        print(f"Árkinyeréskor hiba ennél a Suzuki modellnél: {SuzukiModel}")
        formatted_price = None

    return formatted_price


# Lexus
def getLexusPrices(LexusModel, LexusLink):

    driver.get(LexusLink)

    first_space_index = LexusModel.find(' ')
    ModellNévElsőSzóközig = LexusModel[:first_space_index]

    sleep(3)
    KonfigurátorElemek = driver.find_elements(By.XPATH, '//a[@data-gt-action="click-cta"]')

    for KonfigurátorElem in KonfigurátorElemek:
        if ModellNévElsőSzóközig in KonfigurátorElem.get_attribute('data-gt-label'):
            KeresettModellKonfigurátora = KonfigurátorElem
            break

    KeresettModellKonfigurátora.click()
    sleep(3)

    # Keresse meg az összes olyan elemet, ahol a class tartalmazza az "l-beta-text" szöveget, mert ez az elem veszi fel a modellek konkrét nevét # pl a "https://www.lexus.hu/new-cars/ux/build" oldalon 3 db ilyen elem van: UX 300e, UX 250h AWD, UX 250h FWD
    # MINDEN EGYES LEXUS MODELL ESETÉBEN UGYANEZ A HELYEZ, BÁRMELYIK MODELLNÉL MŰKÖDIK EZ A DINAMIKUS KÓD (AHOGY A TELJES FÜGGVÉNY IS)
    ModellekElemei = driver.find_elements(By.XPATH, '//h2[contains(@class, "l-beta-text")]')

    KeresettModellNév = LexusModel

    # Szűrje az elemeket a keresett modell teljes neve alapján
    for EgyModellEleme in ModellekElemei:
        if KeresettModellNév in EgyModellEleme.text:
            KeresettModellEleme = EgyModellEleme
    try:
        NagySzülő_div = KeresettModellEleme.find_elements(By.XPATH, './../..')
        Ár_div = NagySzülő_div[0].find_elements(By.XPATH, './/div[contains(@class, "CashPrice__Text")]')
        Ár = Ár_div[0].text
        clean_price = ''.join(filter(str.isdigit, Ár))
        formatted_price = f"{clean_price}"
        print("Feldolgozás sikeres")
    except:
        print(f"Hiba történt ezzel a Lexus modellel: {LexusModel}")

    return formatted_price

#Cupra
def getCupraPrices(CupraModel, CupraLink):
    driver.get(CupraLink)
    árLine = 0

    # Sütik elfogadása
    try:
        # Várjunk a gomb betöltésére maximum 10 másodpercig
        button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )

        # Kattintsunk a gombra
        button.click()

    except:
        None


    try:
         # Az árak kinyerése
                price_element = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, f"(//span[translate(normalize-space(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')=translate('{CupraModel}', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz')])")))
                price_element = price_element.find_element(By.XPATH, "..")
                price_element = price_element.find_element(By.XPATH, "..")
                price_element_splitted=price_element.text.split("\n")
                for line in price_element_splitted:
                    if "Ft-tól" in line:
                        árLine = line
                price = árLine
                clean_price = ''.join(filter(str.isdigit, price))
                formatted_price = f"{int(clean_price)}"
                print("Feldolgozás sikeres")
    except:
        print(f"Hiba történt ennél a Cupra modellnél: {CupraModel}")
        formatted_price = None

                
    return formatted_price

# BMW
def getBMWPrices(BMWModel, BMWLink):
    price = None
    formatted_price = None
    clean_price = None
    extracted_chars = []

    driver.get(BMWLink)
    sleep(3)
    

    after_bmw = BMWModel.split('BMW ')[-1]
    first_char = after_bmw[0]
    if first_char == 'i':
        extracted_chars.append('BMWi')
    else:
        extracted_chars.append(first_char)
    
    extracted_chars = []

    after_bmw = BMWModel.split('BMW ')[-1]
    first_char = after_bmw[0]
    if first_char == 'i':
            extracted_chars.append('BMWi')
    else:
            extracted_chars.append(first_char)

    # Sütik elfogadása
    try:
        cookie_banner = driver.execute_script('''
        return document.querySelector('epaas-consent-drawer-shell').shadowRoot.querySelector('button[class="accept-button button-primary"]');
    ''')
        cookie_banner.click()
    except:
        pass
    
    sleep(3)
    for index, char in enumerate(extracted_chars):
        try:
            modellvalasztasbmw = driver.find_element(By.XPATH, f"//a[@class='cmp-modelcard__con-link' and @title='{BMWModel}']")
            modellvalasztasbmw.click()
            sleep(3)  
            
            price = driver.execute_script('''
                let elements = document.querySelector("body > con-app").shadowRoot.querySelector("router-slot > con-configure").shadowRoot.querySelector("con-sales-footer").shadowRoot.querySelector("#salesFooterWrapper > div > div.sales-footer-sections > div.sales-footer-sections-top > div.sales-footer-prices-section > div.sales-footer-total-price-wrapper > button > span");
                return elements && elements.textContent;
            ''')
            
            clean_price = ''.join(filter(str.isdigit, price))
            formatted_price = f"{int(clean_price)}"
            print("Feldolgozás sikeres")
                        
        except:
            try:
                driver.get(BMWLink)
                
                # Sütik elfogadása
                try:
                    cookie_banner = driver.execute_script('''
                    return document.querySelector('epaas-consent-drawer-shell').shadowRoot.querySelector('button[class="accept-button button-primary"]');
                ''')
                    cookie_banner.click()
                except:
                    pass
                
                sleep(3)

                modellvalasztasbmw = driver.find_element(By.XPATH, f"//a[@class='cmp-modelcard__con-link' and @title='{BMWModel}']")
                modellvalasztasbmw.click()
                sleep(20)  
                
                price = driver.execute_script('''
                    let elements = document.querySelector("body > con-app").shadowRoot.querySelector("router-slot > con-configure").shadowRoot.querySelector("con-sales-footer").shadowRoot.querySelector("#salesFooterWrapper > div > div.sales-footer-sections > div.sales-footer-sections-top > div.sales-footer-prices-section > div.sales-footer-total-price-wrapper > button > span");
                    return elements && elements.textContent;
                ''')
                
                clean_price = ''.join(filter(str.isdigit, price))
                formatted_price = f"{int(clean_price)}"
                print("Feldolgozás sikeres")
            except:
                print(f"Error with {BMWModel}")    
    return formatted_price


# Volkswagen
def getVolkswagenPrices(VolkswagenModel, VolkswagenLink):
    
    driver.get(VolkswagenLink)

    formatted_price = 0

    # Sütik elfogadása
    try:
        accept_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )
        accept_button.click()
    except:
        pass 
    
    sleep(5)

    try: 
        chosen_model = driver.find_element(By.XPATH, f"//span[normalize-space()='{VolkswagenModel}']")
        chosen_model = chosen_model.find_element(By.XPATH, "..").find_element(By.XPATH, "..").find_element(By.XPATH, "..").find_element(By.CLASS_NAME, "image-link")
        chosen_model.click()
        sleep(3)
        
        price_element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "(//span[@class='secondary ng-star-inserted'])[1]"))
        )
        
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{clean_price}"
        print("Feldolgozás sikeres")
    
    except:
        try:
            driver.get(VolkswagenLink)

            formatted_price = 0

            # Sütik elfogadása
            try:
                accept_button = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
                )
                accept_button.click()
            except:
                pass 
            
            sleep(5)
            chosen_model = driver.find_element(By.XPATH, f"//span[normalize-space()='{VolkswagenModel}']")
            chosen_model = chosen_model.find_element(By.XPATH, "..").find_element(By.XPATH, "..").find_element(By.XPATH, "..").find_element(By.CLASS_NAME, "image-link")
            chosen_model.click()
            sleep(3)
            
            price_element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "(//span[@class='secondary ng-star-inserted'])[1]"))
            )
            
            price = price_element.text
            clean_price = ''.join(filter(str.isdigit, price))
            formatted_price = f"{clean_price}"
            print("Feldolgozás sikeres")
        except:
            print(f"Hiba történt a következő Volkswagen modell kinyerésénél: {VolkswagenModel}")

    return formatted_price



# Toyota
def GetToyotaPrices(ToyotaModel, ToyotaLink):
    formatted_price = None

    # Lowercase azért, mert az urlekben csak ilyen betűk vannak, és csak ilyen urlekkel tudjuk megnyitni a weboldalakat
    ToyotaModel = ToyotaModel.lower()

    # Számoljuk meg, hogy szóköz van a beolvasott modellnévben
    SzóközSzám = ToyotaModel.count(' ')

    first_space_index = ToyotaModel.find(' ')
    second_space_index = ToyotaModel.find(' ', first_space_index + 1)

    ModellNévElsőSzóközig = ToyotaModel[:first_space_index].lower()
    ModellnévMásodikSzóközig = ToyotaModel[:second_space_index].lower()

    # MINDEN LEHETSÉGES MODELLNÉV ESET DINAMIKUSAN LEKEZELVE:

    if SzóközSzám == 0: #Például Excelből beolvasott modellnév: Yaris, weboldal: https://www.toyota.hu/modellek/yaris
        ToyotaTeljesUrl = ToyotaLink+"/"+ToyotaModel.lower()
        #print(ToyotaTeljesUrl, "0szóköz")
        driver.get(ToyotaTeljesUrl)


    if SzóközSzám == 1: #Például Excelből beolvasott modellnév: Prius Plug-in, weboldal: https://www.toyota.hu/modellek/prius-plugin
        ToyotaTeljesUrl = ToyotaLink+"/"+ToyotaModel.replace('-', "").replace(' ', '-')
        #print(ToyotaTeljesUrl, "1szóköz teljesszöveg")
        driver.get(ToyotaTeljesUrl)

        # A Toyota hivatalos oldala úgy van beállítva, hogy ha nem létező urlt próbálunk megnyitni (pl. https://www.toyota.hu/modellek/camry-hybrid),
        # akkor automatikusan átirányít minket (a drivert is természetesen) az Árlisták oldalra. Tehát ha a driver jelenlegi url-je(driver.current_url)
        # nem egyezik azzal az url-lel amellyel az előbb megnyitottuk (driver.get(ToyotaTeljesUrl), akkor biztos, hogy nem létező url-t próbáltunk megnyitni a driverrel
        
        if driver.current_url != ToyotaTeljesUrl: #Például Excelből beolvasott modellnév: Camry Hybrid, weboldal: https://www.toyota.hu/modellek/camry  
            ToyotaTeljesUrl = ToyotaLink+"/"+ModellNévElsőSzóközig
            driver.get(ToyotaTeljesUrl)
            #print(ToyotaTeljesUrl, "1szóköz elsőszóközig")

    if SzóközSzám >= 2: #Például Excelből beolvasott modellnév: Proace City Verso EV, weboldal: https://www.toyota.hu/modellek/proace-city-verso-ev
        ToyotaTeljesUrl = ToyotaLink+"/"+ToyotaModel.replace('-', "").replace(' ', '-')
        #print(ToyotaTeljesUrl, "többszóköz teljesszöveg")
        driver.get(ToyotaTeljesUrl)

        if driver.current_url != ToyotaTeljesUrl: #Például Excelből beolvasott modellnév: Corolla Sedan 1.8 Hybrid, weboldal: https://www.toyota.hu/modellek/corolla-sedan
            ToyotaTeljesUrl = ToyotaLink+"/"+ModellnévMásodikSzóközig.replace('-', "").replace(' ', '-') 
            #print(ToyotaTeljesUrl, "többszóköz másodikszóközig")
            driver.get(ToyotaTeljesUrl)
            
            if driver.current_url != ToyotaTeljesUrl: #Például Excelből beolvasott modellnév: C-HR 1.8 Hybrid, weboldal: https://www.toyota.hu/modellek/c-hr
                ToyotaTeljesUrl = ToyotaLink+"/"+ModellNévElsőSzóközig 
                driver.get(ToyotaTeljesUrl)
                #print(ToyotaTeljesUrl, "többszóköz elsőszóközig")
    
    # Generalizálható rendszert szeretnénk fejleszteni (hardcoding elkerülése):
    # Az alábbi megoldási módszer ebben a leghatékonyabb, amely a 29 db hivatalosan elérhető Toyota modell közül 5 db modell kivételével bármelyik
    # árát képes visszaadni csupán az input excelben a lekérdezni kívánt új Toyota modell sorának hozzáadásával
    # a 29-ből az 5 db modellnél amelyiknél nem működik, ott többnyire azért nem működik, mert ezek nagy része új = csak előfoglalható (nem elérhető még) = csak előzetes árlista pdf-ek, melyek html és pdf struktúrája nagyon eltér a többi 24 db ugyanolyantól
    # az érintett 5 db modell: Új Yaris, ÚJ Toyota C-HR, ÚJ Land Cruiser, GR Yaris, Yaris VAN

    # A többi 24 db modellnél:
    # Ezek közül számos modell hivatalos weboldalán nem található ár (pl. van "Indulóár (Tartalmazza az ÁFA-t)" - https://www.toyota.hu/modellek/corolla-sedan VS nincs ár -https://www.toyota.hu/modellek/gr86,),
    # azonban a modell weboldalakon található hivatalos Árlista pdf-ekben MIND a 24 db modell esetében megtalálható az ár (pl. https://pdf.sites.toyota.hu/arlista_toyota_c-hr_23.pdf)
    # ezt a pdf url-t minden oldalon annak az elemnek a href attribútum értéke tartalmazza, amely data-gt-action attribútuma = "pricelistdownload"

    # Keresés az elemre a 'data-gt-action="pricelistdownload"' attribútum alapján, mert bármelyik Toyota modell weboldalán ez az element tartalmazza az árlista pdf linkjét
    sleep(3)
    target_element = driver.find_element(By.XPATH, '//*[@data-gt-action="pricelistdownload"]')

    # Az elem href attribútum értékének kinyerése, azaz az Árlista pdf linkjének kinyerése
    pdf_link = target_element.get_attribute('href')

    response = requests.get(pdf_link)
    response.raise_for_status()

    pdf_file = io.BytesIO(response.content)

    # Megnyitjuk a PDF-et pdfplumber segítségével
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            break

    # Szétbontjuk a pdf-ről kinyert szöveget sorokra
    lines = text.splitlines()

    # Megkeressük azt a sort, amely "tól"-t tartalmaz:
    # BÁRMELYIK hivatalos Toyota Árlista pdf-ben a "Ft-tól"-t tartalmazó "sorban" található a modell ára, példa: https://pdf.sites.toyota.hu/arlista_corolla_sedan_2024.pdf,
    # Azonban a "-" karakter keresésével gyakran adódik probléma (The character U+2011 "‑" could be confused with the ASCII character U+002d "-", which is more common in source code.),
    # ezért csak a "tól" szövegrészletet keressük
    for line in lines:
        if "tól" in line:
            árLine = line
            break

    # Ebből a sorból kivesszük a csak számokat és szóközöket tartalmazó szövegrészeket
    numbers_text = re.sub(r'[^\d\s]', '', árLine)

    # Ahol kettő vagy több szóköz van ebben a szövegben, ott felbontjuk (splitteljük) a szöveget
    # Azért kettő vagy több szóköz, mert ahol csak 1 szóköz van a számjegyek között, az csak az erzescsoportok közti tagolást jelenti 1-1 számnál, pl: "10 260 000  700 000" ( a pdf struktúrában a szövegkinyerésnél ez a 2 szám egy sorban van)
    numbers_str = re.split(r'\s{2,}', numbers_text)

    # Az egyes számok átalakítása egész számokká
    numbers = [int(n.replace(' ', '')) for n in numbers_str]

    # A két szám közül a nagyobbik kiválasztása, 
    # Mivel a kisebbik szám mindig, minden egyes Toyota modell esetében biztosan az árkedvezményt fogja jelenti, míg a nagyobbik szám a modell árát, egy példa: "https://pdf.sites.toyota.hu/arlista_corolla_sedan_2024.pdf"
    try:
        formatted_price = max(numbers)
        print("Feldolgozás sikeres")
    except:
        print(f"Hiba történt ezzel a Toyota modellel: {ToyotaModel}")
    
    return formatted_price
  

# Volvo
def GetVolvoPrices(VolvoModel, VolvoLink):
    SzóközSzám = VolvoModel.count(' ')

    # Ha nincs szóköz a modell nevében, akkor a teljes szöveget visszaadjuk
    if SzóközSzám == 0:
        konkrét_modell_név_urlbe = VolvoModel

    # Ha van szóköz a modell nevében
    else:
        # Ha 1 db szóköz van benne, akkor az biztosan, BÁRMELYIK modell esetében így fog kinézni: '{Modellnév} Recharge',
        # és mivel a Recharge-t - ami Volvo terminológiával azt jelenti, hogy az autó Plug-in hybrid vagy Tisztán elektromos - külön kezeljük a 'Meghajtás (Hyundai, Volvo)' oszlopban (PHEV vagy BEV) egy "-hybrid" vagy "-electric" tag hozzáadásával az url-be,
        # ezért 1 db szóköz esetén csak a szóköz előtti modellnévre van szükségünk a megfelelő url-hez
        # pl Modell neve: 'S90 Recharge', weboldal: https://www.volvocars.com/hu/cars/s90-hybrid/
        if SzóközSzám == 1:
            # Megkeressük az első szóközt
            first_space_index = VolvoModel.find(' ')

            # Visszaadjuk a szöveget az első szóközig
            konkrét_modell_név_urlbe = VolvoModel[:first_space_index]
        
        # Ha több szóköz van a modell nevében, az csak a 'V90 Cross Country' és 'V60 Cross Country' modellek esetében fordulhat elő,
        # ekkor a teljes modellnévre szükségünk van a linkbe
        # pl. Modell neve: 'V60 Cross Country', weboldal: 'https://www.volvocars.com/hu/v/cars/v60-cross-country/shop', ezért a szóközök helyett kötőjelre lesz szükségünk az urlbe
        else:
            konkrét_modell_név_urlbe = VolvoModel.replace(" ", "-")


    konkrét_modell_tipus = df["Meghajtás (Hyundai, Volvo)"].iloc[index,]

    if konkrét_modell_tipus == 'PHEV':
        konkrét_modell_tipus_urlbe = "-hybrid"

    # Egy Volvo modell ha hibrid, akkor csak Plug-in hybrid (PHEV) lehet, így ez a rész felesleges és zavaró, ezért komment 
    # if konkrét_modell_tipus == 'HEV':
    #     konkrét_modell_tipus_urlbe = "-hybrid"
    
    if konkrét_modell_tipus == 'BEV':
        konkrét_modell_tipus_urlbe = "-electric"

    if konkrét_modell_tipus == 'MHEV':
        konkrét_modell_tipus_urlbe = ""


    def extract_text_after_prefix(text, prefix="Kezdőár"):
        # Ellenőrizzük, hogy a szöveg a megadott előtaggal kezdődik-e
        if text.startswith(prefix):
            # Visszaadjuk a szöveget az előtag után, eltávolítva a felesleges részeket
            cleaned_text = text[len(prefix):].strip().replace(" Ft", "").replace(" ", "")
            return int(cleaned_text) if cleaned_text.isdigit() else cleaned_text

        # Ha nincs az előtag, a teljes szöveget tisztítjuk és számmá alakítjuk
        cleaned_text = text.replace(" Ft", "").replace(" ", "")
        return int(cleaned_text) if cleaned_text.isdigit() else cleaned_text


    VolvoLinkModel = VolvoLink+"/cars/"+konkrét_modell_név_urlbe+konkrét_modell_tipus_urlbe
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(VolvoLinkModel)

    # Sütik elfogadása
    try:
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Elfogadás')]"))
        ).click()
    except:
        pass

    price = None

    # 'Price' elem megtalálása a weboldalon ('div.ac' azonosítja)
    try:
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "div.ac > small"))
        )
        price_element = driver.find_element(By.CSS_SELECTOR, "div.ac > small")

        price = extract_text_after_prefix(price_element.text)
        print("Feldolgozás sikeres")
        
    except:
            driver = webdriver.Chrome()
            driver.maximize_window()
            VolvoLinkModelShop = VolvoLink+"/v/cars/"+konkrét_modell_név_urlbe+konkrét_modell_tipus_urlbe+"/shop"
            driver.get(VolvoLinkModelShop)

            try:
                WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Elfogadás')]"))
                ).click()
            except:
                pass

            try:
                #Az árat tartalmazó elem megtalálása
                price_element = driver.find_element(By.XPATH, "//em[@data-autoid='imageWithText:subTitle']")

                price = extract_text_after_prefix(price_element.text)
                print("Feldolgozás sikeres")
            except:
                print(f"Hiba történt ezzel a Volvo modellel: {VolvoModel}")
        
    return price


# MG
def getMGPrices(MGModel, MGLink):
        MGModelUrlbe = MGModel.replace(" ", "_")
  
        driver.get(MGLink+"?model="+MGModelUrlbe)
        # '?model=' azért kell bele, mert egy MG modell konfigurátor linkje minden esetben, minden modellnél a következőképpen néz ki (az 'MG_EHS' rész helyett mindig az aktuális modell név):
        # "https://mggaal.hu/konfigurator?model=MG_EHS", ahol a '?model=' (statikus, nem változó) részen kívül mindent (konfigurátor linkje, aktuális modell neve) az input excel fájlból olvasunk be

        ÁrElementek = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.XPATH, "//span[contains(text(), 'Ft')]"))
        )

        # A legtöbb MG modell esetében 2 ár van feltüntetve a konfigurátor weboldalán, ekkor az egyik MINDIG, BÁRMELYIK MODELL ESETÉBEN át van húzva (elavult ár), míg a másik nincs áthúzva (valós ár)
        # Ennél a résznél a nem áthúzott árat kapjuk vissza
        for ÁrElement in ÁrElementek:
            szülőElement = ÁrElement.find_element(By.XPATH, '..') # A számot (árat) tartalmazó element szülő span-jének a style attribútumában van beállítva az áthúzás mint formázás
            if 'line-through' not in szülőElement.get_attribute('style'): # Ha nincs áthúzva az ár, akkor a számot (árat) tartalmazó element szülő span-jének nincs megadva a style attribútuma (üres string) BÁRMELYIK MODELL ESETÉBEN
                ValósÁr = ÁrElement.text
        try:     
            clean_price = ''.join(filter(str.isdigit, ValósÁr))
            formatted_price = f"{clean_price}"
            print("Feldolgozás sikeres")
        except:
            print(f"Hiba történt ezzel a modellel: {MGModel}")

    
        return formatted_price


####################################################################################
print("A program elkezdett futni, kérjük várjon")

driver = webdriver.Chrome()
driver.maximize_window()

ExcelFileBeolvasásra = 'SAP_arak_lekerdezese.xlsx' 
df = pd.read_excel(ExcelFileBeolvasásra, header=0)

modellek_neve = df['Modell neve (valós)']
márkák_neve = df['Márka']

wb = load_workbook(ExcelFileBeolvasásra)
ws = wb['Lekérdezés']
wsMárkaURLek = wb['Márka-URL']

# Keresse meg az első szabad oszlopot (végigmegyünk az oszlopokon az 'A' oszloptól kezdve, keresve az első szabadot)
szabad_oszlop = 'A'
while ws[szabad_oszlop + '1'].value is not None:
    szabad_oszlop = chr(ord(szabad_oszlop) + 1)  # Lépés a következő oszlopra

BeolvasottModellekSzáma = modellek_neve.count()

# tqdm: haladásjelző
for index, konkrét_márka_név in tqdm(enumerate(márkák_neve), total=BeolvasottModellekSzáma, desc="Haladásjelző", unit = 'modell'):
    # 2 Szükséges üres sor az output igényessége miatt
    print()
    print()

    konkrét_márka_név = konkrét_márka_név.lower()

    if konkrét_márka_név == 'audi':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"Az Audi {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Audi':
                AudiLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = (getAudiPrices(konkrét_modell_név, AudiLink))

    elif konkrét_márka_név == 'peugeot':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        print(f"A {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Peugeot':
                PeugeotLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = (getPeugeotPrices(PeugeotLink))

    elif konkrét_márka_név == 'bmw':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        print(f"A {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'BMW':
                BMWLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = (getBMWPrices(konkrét_modell_név, BMWLink))

    elif konkrét_márka_név == 'kia':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Kia {konkrét_modell_név} modell feldolgozása folyamatban")
        
        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Kia':
                KiaLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = (getKiaPrices(konkrét_modell_név, KiaLink))

    elif konkrét_márka_név == 'suzuki':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Suzuki {konkrét_modell_név} modell feldolgozása folyamatban")
        
        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Suzuki':
                SuzukiLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = (getSuzukiPrices(konkrét_modell_név, SuzukiLink))

    elif konkrét_márka_név == 'cupra':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        print(f"A {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Cupra':
                CupraLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = getCupraPrices(konkrét_modell_név, CupraLink)

    elif konkrét_márka_név == 'volkswagen':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Volkswagen {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Volkswagen':
                VolkswagenLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = getVolkswagenPrices(konkrét_modell_név, VolkswagenLink)

    elif konkrét_márka_név == 'mg':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        print(f"Az {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'MG':
                MGLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = getMGPrices(konkrét_modell_név, MGLink)

    elif konkrét_márka_név == 'hyundai':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Hyundai {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Hyundai':
                HyundaiLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = getHyundaiprices(konkrét_modell_név, HyundaiLink)

    elif konkrét_márka_név == 'toyota':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Toyota {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Toyota':
                ToyotaLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = GetToyotaPrices(konkrét_modell_név, ToyotaLink)

    elif konkrét_márka_név == 'skoda':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Skoda {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Skoda':
                SkodaLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = getSkodaPrices(konkrét_modell_név, SkodaLink)

    elif konkrét_márka_név == 'lexus':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Lexus {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Lexus':
                LexusLink = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = getLexusPrices(konkrét_modell_név, LexusLink)

    elif konkrét_márka_név == 'volvo':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        print(f"A Volvo {konkrét_modell_név} modell feldolgozása folyamatban")

        for row in range(1, wsMárkaURLek.max_row + 1):
            if wsMárkaURLek[f'A{row}'].value == 'Volvo':
                VolvoLinkModel = wsMárkaURLek[f'B{row}'].value
                break

        cell = f'{szabad_oszlop}%d'  % (index + 2)
        ws[cell] = GetVolvoPrices(konkrét_modell_név, VolvoLinkModel)

else:
    print()
    print("Az Excel fájl feldolgozása sikeresen megtörtént, bővítése folyamatban. Kérjük még ne nyissa meg a fájlt.")


# A lekérdezett árak formázása (lekérdezés oszlopának celláinak formázása az első sor kihagyásával)
for row in range(2, ws.max_row + 1):  # Kezdve a 2. sortól (azaz a G2 cellától)
    cell = ws[f'{szabad_oszlop}{row}']
    if cell.value is not None:  # Ellenőrzés, hogy a cella értéke nem None
        try:
 
            cell.value = int(cell.value)
            cell.number_format = '#,##0 Ft'
        except ValueError:
            # Ha a cella értéke nem alakítható át számmá, hagyja figyelmen kívül
            pass

# Első sor a lekérdezés dátuma
current_datetime = datetime.now()
new_column_name = f'{current_datetime}'
ws[f'{szabad_oszlop}1'] = current_datetime

# Lekérdezés oszlopának szélességbeállítása
ws.column_dimensions[szabad_oszlop].width = 20

# Változtatások mentése
wb.save(ExcelFileBeolvasásra) 
wb.close()

driver.quit()

print("Az Excel fájl sikeresen bővült az imént lekérdezett árakkal, most már megnyithatja a fájlt. A program leáll.")